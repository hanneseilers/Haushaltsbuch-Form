import openpyxl as xl
import openpyxl.worksheet.worksheet as xl_worksheet
import openpyxl.cell.cell as xl_cell
import easygui as gui
from os import access, R_OK, W_OK
from os.path import exists, isfile
from shutil import copyfile
from datetime import date
import config as cfg



def process() -> None:
    # init year and date values
    year = date.today().year
    year = str(year)[-2] + str(year)[-1]
    month = str(date.today().month).zfill(2)
    day = str(date.today().day).zfill(2)

    # ask if to start or to end
    title_start = "Continue?"
    msg_start = "continue form?"
    while gui.ynbox(title=title_start, msg=msg_start):

        # ask for date
        values_date = ask_for_date(year, day, month)
        if not values_date:
            continue
        year, day, month = values_date

        # get workbook file (or create it)
        wb_data = get_workbook(year, month)

        # check if workbook is readable
        if not wb_data:
            title_error = "ERROR"
            msg_error = f"error reading file {month}/{year}"
            gui.msgbox(title=title_error, msg=msg_error)
            continue
        workbook_filename, workbook, active_worksheet = wb_data

        # read header
        line_data = get_workbook_line_data(worksheet=active_worksheet, row=cfg.header_row, column_start=cfg.header_column_start,
                                                column_end=cfg.header_column_end)
        print(f"loaded {len(line_data)} header columns")

        # starting entering actual data
        if len(line_data) > 0:
            sections = group_sections_by_bg_color(line_data)

            # ask for section
            section_selection = ask_for_section(sections)
            if not section_selection:
                continue
            choice, section = section_selection
            print(f"selected section with color {str(choice)}")

            # ask for category
            category_data = ask_for_category(section)
            if not category_data:
                continue
            category_number, category_name, category = category_data

            # ask for value
            value = ask_for_value()
            if not value:
                continue

            # calculate day row
            if add_value_to_worksheet(worksheet=active_worksheet,
                                      day=day,
                                      category=category,
                                      value=value):
                workbook.save(workbook_filename)
                print(f"data saved to {workbook_filename}")
            else:
                print(f"data NOT saved to {workbook_filename}")

def ask_for_date(year:str, day:str, month:str) -> tuple[str, str, str]:
    # ask for date
    title_date = "Date?"
    msg_date = "enter date:"
    fields_date = ["year", "day", "month"]
    values_date = [year, day, month]
    year, day, month = gui.multenterbox(title=title_date, msg=msg_date, fields=fields_date, values=values_date)
    year = str(year).zfill(2)
    month = str(month).zfill(2)
    day = str(day).zfill(2)

    return year, day, month


def __ask_for_section_choice(sections:dict|list,
                             title:str= "Select Section",
                             msg:str= "select the section of your data entry",
                             use_gui_buttons:bool=True) -> str:
    choices = get_choices(data=sections, use_choice_value=isinstance(sections, list))

    if use_gui_buttons:
        choice = gui.buttonbox(title=title, msg=msg, choices=choices)
    else:
        choice = gui.choicebox(title=title, msg=msg, choices=choices)

    return str(choice).strip()


def ask_for_section(sections:dict) -> tuple[str, list] | None:
    choice = __ask_for_section_choice(sections)
    choice = choice.replace('[', '').replace(']', '').split(":")[1].strip()
    section = sections[choice]

    if not section:
        return None

    print(f"selected section {str(choice)}")
    return choice, section


def ask_for_category(section:list) -> tuple[int, str, xl_cell.Cell]|None:
    category_name = __ask_for_section_choice(sections=section, title="Select Column",
                                             msg="select category where to enter value",
                                             use_gui_buttons=False)

    if not category_name:
        return None

    category_number = str(category_name).split(":")[0].strip()
    category_number = int(category_number) - 1
    category = section[category_number]

    print(f"selected category #{category_number}: {str(category.value)}")
    return category_number, category_name, category


def ask_for_value(category_name:str = "category") -> str|None:
        title_value = "Data Value"
        msg_value = f"enter value to add for\n{category_name}"
        value = gui.enterbox(title=title_value, msg=msg_value)

        if not value:
            return None

        value = str(value).strip().replace(',', '.')

        print(f"received value = {str(value)}")
        return value


def get_workbook_filename(base_file: str, year: str = "", month: str = "") -> str:
    return base_file.replace("%y", str(year)).replace("%d", str(month))


def get_workbook(year:str, month:str) -> tuple[str, xl.Workbook, xl_worksheet.Worksheet]| None:
    # get workbook file (or create it)
    wb_filename = get_workbook_filename(base_file=cfg.monthly_filename, year=year, month=month)
    if not (
            exists(wb_filename) or
            isfile(wb_filename)
    ):
        wb_base_filename = get_workbook_filename(base_file=cfg.base_filename)
        print(f"copying file {wb_base_filename} to {wb_filename}")
        copyfile(wb_base_filename, wb_filename)

    # check if workbook is readable
    if not (
            access(wb_filename, R_OK) or
            access(wb_filename, W_OK)
    ):
        return None

    # load header
    wb = xl.load_workbook(wb_filename, keep_vba=True, keep_links=True, )

    # load all sheets
    if isinstance(wb, xl.Workbook):
        for ws_name in wb.sheetnames:
            print(f"loading worksheet {ws_name}")
            _ws = wb[ws_name]


    return wb_filename, wb, wb.active


def get_workbook_line_data(worksheet: xl.Workbook, row=1, column_start=2, column_end=100) -> list:
    header = []
    for column in range(column_start, column_end + 1):
        cell = worksheet.cell(row=row, column=column)
        if cell.value and len(str(cell.value).strip()) > 0:
            header.append(cell)
    return header


def group_sections_by_bg_color(line_data:list) -> dict[str, list]:
    sections = {}

    # group columns into sections by color
    for cell in line_data:

        # get section color, if possible
        color = "default"
        if hasattr(cell, "fill") \
                and hasattr(cell.fill, "bgColor") \
                and hasattr(cell.fill.bgColor, "rgb"):

            if isinstance(cell.fill.bgColor.rgb, str):
                color = cell.fill.bgColor.rgb

        # create section entry for new color, if needed
        if color not in sections.keys():
            sections[color] = []
        sections[color].append(cell)

    print(f"found {len(sections)} sections ({str(sections.keys())})")
    return sections


def get_choices(data:dict|list, use_choice_value:bool = False) -> list[str]:
    choices = []
    _n = 1

    if isinstance(data, dict):
        data = data.keys()

    for _choice in data:
        if use_choice_value:
            choices.append(f"{str(_n)}: {str(_choice.value)}")
        else:
            choices.append(f"{str(_n)}: {str(_choice)}")
        _n += 1

    return choices


def add_value_to_worksheet(worksheet: xl_worksheet.Worksheet,
                           day: int | str,
                           category:xl_cell.Cell, value: str | int) -> xl.Workbook | None:
    # calculate day row
    data_cell_row = cfg.date_row_start + int(day) - 1
    if data_cell_row <= cfg.date_row_end:

        # calculate cell
        data_cell_value = category.column_letter + str(data_cell_row)
        print(f"using data cell {str(data_cell_value)}")

        # add value
        data_value = worksheet[data_cell_value].value
        if not data_value:
            data_value = "="

        data_value = str(data_value)
        if "=" not in data_value:
            data_value = "=" + data_value

        data_value += f"+{str(value)}"
        print(f"new value for {data_cell_value} is '{data_value}'")
        worksheet[data_cell_value].value = data_value

        return worksheet

    return None


# ---- MAIN ----
if __name__ == '__main__':
    process()
